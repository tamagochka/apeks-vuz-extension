import logging
import os
from typing import TypeAlias

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.page import PageMargins
from ..core.reports.ExcelStyles import ExcelStyle

from config import FlaskConfig, MongoDBSettings
from ..core.services.staff_various_document_service import StaffVariousGroupDocStructure

Filename: TypeAlias = str


def get_various_report_data(document: dict, faculty_names: dict):
    daytime = document.get("daytime")
    result = {
        "date": document.get("date"),
        "daytime": MongoDBSettings.DAYTIME_NAME.get(
            daytime, f"неверный указатель времени - {daytime}"
        ),
        "faculty_data": {},
        "busy_types": [],
        "illness_types": [],
    }
    for group in document.get("groups").values():
        group_data = StaffVariousGroupDocStructure(**group)
        for absence in group_data.absence:
            if absence not in result["busy_types"]:
                result["busy_types"].append(absence)
        for illness in group_data.absence_illness:
            if illness not in result["illness_types"]:
                result["illness_types"].append(illness)
        faculty_data = result["faculty_data"].setdefault(group_data.faculty, {})
        courses_data = faculty_data.setdefault('courses_data', {})
        course_data = courses_data.setdefault(
            group_data.course, {"total": 0, "absence": {}, "absence_illness": {}}
        )
        course_data["total"] += group_data.total
        for absence_name, students in group_data.absence.items():
            course_data["absence"].setdefault(absence_name, 0)
            course_data["absence"][absence_name] += len(students)
        for illness_name, students in group_data.absence_illness.items():
            course_data["absence_illness"].setdefault(illness_name, 0)
            course_data["absence_illness"][illness_name] += len(students)

    def sort_faculty(faculty_name):
        if faculty_name in faculty_names:
            return faculty_names[faculty_name]
        return max(faculty_names.values()) + 1

    result["faculty_data"] = {
        faculty: result["faculty_data"][faculty]
        for faculty in sorted(result["faculty_data"], key=sort_faculty)
    }
    for faculty, courses_data in result["faculty_data"].items():
        result["faculty_data"][faculty] = {
            course: courses_data[course] for course in sorted(courses_data)
        }
    for faculty in result['faculty_data']:
        result['faculty_data'][faculty]['branch_id'] = faculty_names[faculty][1]
    return result


def generate_various_staff_report(
    various_report_data,
    busy_types: dict,
    illness_types: dict,
    branches: dict
) -> Filename:
    """
    Формирует отчет - строевая записка постоянного состава.

    Parameters
    ----------
        various_report_data
            данные о наличии личного состава
        busy_types
            данные о видах отвлечений
        illness_types
            данные о видах болезней

    Returns
    -------
        str
            название файла
    """

    workdate = various_report_data.get("date")
    title = (
        f"Строевая записка переменного состава за {workdate} "
        f"по состоянию на {various_report_data.get('daytime').lower()}"
    )
    wb = Workbook()
    ws = wb.active
    ws.title = workdate
    ws.page_setup.orientation = "landscape"
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.5, bottom=0.2, header=0.2, footer=0.2
    )
    ws.page_setup.fitToWidth = 1

    def write_cell_data(cell_row, cell_col, val, style=ExcelStyle.Number, grey=False):
        """Записывает данные в ячейку."""
        ws.cell(cell_row, cell_col).value = val
        ws.cell(cell_row, cell_col).style = style
        if grey:
            ws.cell(cell_row, cell_col).fill = ExcelStyle.GreyFill
        ws.cell(cell_row, cell_col).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
            shrink_to_fit=True,
        )

    # Заголовок таблицы
    row = 1
    col = 1
    ws.cell(row, col).value = title
    ws.cell(row, col).style = ExcelStyle.Header
    ws.cell(row, col).font = Font(name="Times New Roman", size=14, bold=True)
    ws.cell(row, col).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True
    )

    # Заголовки для первой колонки
    header_names = {
        "branches": "Филиалы",
        "faculties": "Факультеты",
        "courses": "Курсы",
        "total": "По списку",
        "in_stock": "В строю",
        "absence": "Отсутствуют",
    }
    for item in various_report_data.get("busy_types"):
        header_names[item] = busy_types.get(item, item)
    header_names["illness"] = "Больны"
    for item in various_report_data.get("illness_types"):
        header_names[item] = illness_types.get(item, item)

    # Номера строк дла заголовков
    header_rows = {header: h_row for h_row, header in enumerate(header_names, start=2)}
    bold_rows = ["branches", "faculties", "courses", "total", "in_stock", "absence", "illness"]
    for header in header_names:
        row = header_rows[header]
        style = ExcelStyle.BaseBold if header in bold_rows else ExcelStyle.Base
        write_cell_data(row, col, header_names[header], style)
    ws.column_dimensions[get_column_letter(col)].width = 34

    # Факультеты
    col = 2
    faculty_data = various_report_data.get("faculty_data")
    for branch in branches:
        # Список итоговых столбцов для исключения из суммы
        exclude_total_col = []
        start_col = col
        count_facs = 0
        for faculty, courses in faculty_data.items():
            if str(courses['branch_id']) != str(branch):
                continue
            else:
                count_facs += 1
            write_cell_data(
                header_rows["faculties"], col, faculty, style=ExcelStyle.BaseBold
            )
            # Курсы
            for course, course_data in courses['courses_data'].items():
                row = header_rows["courses"]
                ws.column_dimensions[get_column_letter(col)].width = 4
                write_cell_data(row, col, f"{course}", style=ExcelStyle.BaseBold)
                write_cell_data(header_rows["total"], col, course_data.get("total"))
                for absence, value in course_data["absence"].items():
                    write_cell_data(header_rows[absence], col, value)
                for illness, value in course_data["absence_illness"].items():
                    write_cell_data(header_rows[illness], col, value)

                # Расчет больных
                letter = get_column_letter(col)
                row = header_rows["illness"]
                total_illness = len(course_data["absence_illness"])
                formula = f"=SUM({letter}{row + 1}:{letter}{row + total_illness})"
                write_cell_data(row, col, formula)

                # Расчет отсутствующих
                row = header_rows["absence"]
                total_absence = len(course_data["absence"])
                formula = f"=SUM({letter}{row + 1}:{letter}{row + total_absence + 1})"
                write_cell_data(row, col, formula)

                # Расчет находящихся в строю
                row = header_rows["in_stock"]
                formula = f"={letter}{row - 1} - {letter}{row + 1}"
                write_cell_data(row, col, formula)

                col += 1

            for header in header_rows:
                write_cell_data(
                    header_rows["courses"], col, "Итого", style=ExcelStyle.BaseBold
                )
                letter_from = get_column_letter(col - len(courses['courses_data']))
                letter_to = get_column_letter(col - 1)
                ws.column_dimensions[get_column_letter(col)].width = 7
                if header not in ("courses", "faculties", "branches"):
                    row = header_rows[header]
                    value = f"=SUM({letter_from}{row}:{letter_to}{row})"
                    write_cell_data(row, col, value, style=ExcelStyle.BaseBold)
            exclude_total_col.append(col)
            col += 1

            ws.merge_cells(
                start_row=header_rows["faculties"],
                start_column=col - len(courses['courses_data']) - 1,
                end_row=header_rows["faculties"],
                end_column=col - 1,
            )

        end_col = col

        if count_facs > 0:  # если в филиале есть заполненные данные по факультетам
            # Добавление итоговой графы
            write_cell_data(
                header_rows["faculties"], col, "Итого по филиалу", style=ExcelStyle.BaseBold
            )
            ws.merge_cells(
                start_row=header_rows["faculties"],
                start_column=col,
                end_row=header_rows["courses"],
                end_column=col,
            )
            # добавляем заголовок филиала
            write_cell_data(header_rows["branches"], start_col, branches[branch], style=ExcelStyle.BaseBold)
            
            
            ws.cell(header_rows["branches"], start_col).font = Font(color='FFFFFF', bold=True)


            ws.cell(header_rows["branches"], start_col).alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
            # объединяем ячейки заголовка филиалов
            ws.merge_cells(
                start_row=header_rows["branches"],
                start_column=start_col,
                end_row=header_rows["branches"],
                end_column=end_col
            )
            ws.column_dimensions[get_column_letter(col)].width = 11
            for header in header_rows:
                letter_from = get_column_letter(start_col)
                letter_to = get_column_letter(end_col - 1)
                if header not in ("courses", "faculties", "branches"):
                    row = header_rows[header]
                    formula = f"=SUM({letter_from}{row}:{letter_to}{row})"
                    for excl_col in exclude_total_col:
                        formula += f"-{get_column_letter(excl_col)}{row}"
                    write_cell_data(row, col, formula, style=ExcelStyle.BaseBold)
            col += 1
            # обводим филиал толстой линией
            ExcelStyle.set_border(ws, (f"{get_column_letter(start_col)}{header_rows['branches']}:"
                            f"{get_column_letter(end_col)}{len(header_names) + 1}"))

    # Объединение ячеек заголовка
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col - 1)

    # Выделение серым заголовка
    for row in (header_rows["faculties"], header_rows["courses"]):
        for col in range(1, col):
            ws.cell(row, col).fill = ExcelStyle.GreyFill

    # Выделение темно серым строки с филиалами
    for col in range(1, end_col):
        ws.cell(header_rows['branches'], col).fill = ExcelStyle.DarkFill

    # Формируем файл отчета
    filename = f"{title}.xlsx"
    wb.save(os.path.join(FlaskConfig.EXPORT_FILE_DIR, filename))
    logging.debug(f"Файл '{filename}' успешно сформирован")
    return filename
