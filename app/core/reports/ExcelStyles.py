from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

class ExcelStyle(object):
    """Стили оформления для экспортируемых файлов Excel."""

    ThinBorder = Side(style="thin", color="000000")
    ThickBorder = Side(style="thick", color="000000")
    AllBorder = Border(
        left=ThinBorder, right=ThinBorder, top=ThinBorder, bottom=ThinBorder
    )

    Header = NamedStyle(name="header")
    Header.font = Font(name="Times New Roman", size=12, bold=True)
    Header.border = AllBorder
    Header.alignment = Alignment(wrap_text=True)

    HeaderSmall = NamedStyle(name="header_small")
    HeaderSmall.font = Font(name="Times New Roman", size=10, bold=True)
    HeaderSmall.border = AllBorder
    HeaderSmall.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
                shrink_to_fit=True,
            )

    Base = NamedStyle(name="base")
    Base.font = Font(name="Times New Roman", size=11)
    Base.border = AllBorder
    Base.alignment = Alignment(wrap_text=True)

    Base_No_Wrap = NamedStyle(name="base_no_wrap")
    Base_No_Wrap.font = Font(name="Times New Roman", size=11)
    Base_No_Wrap.border = AllBorder
    # Base.alignment = Alignment(wrap_text=True)

    Number = NamedStyle(name="number")
    Number.font = Font(name="Times New Roman", size=9)
    Number.border = AllBorder
    Number.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True
    )

    BaseBold = NamedStyle(name="base_bold")
    BaseBold.font = Font(name="Times New Roman", size=11, bold=True)
    BaseBold.border = AllBorder
    BaseBold.alignment = Alignment(wrap_text=True)

    GreyFill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    DarkFill = PatternFill(start_color="888888", end_color="888888", fill_type="solid")
    
    def set_border(ws: Worksheet, cells: str):
        """
        Обвести группу ячеек толстой линией.

        Parameters
        ----------
            ws
                рабочий лист книги Excel на котором произовдися операция
            cells
                координаты верхнего левого и павого нижнего углов группы ячеек,
                которые необходимо обвести толстой линией
                в формате "A1:C3"
        """

        rows = ws[cells]
        for row in rows:
            row[0].border = Border(
                left=ExcelStyle.ThickBorder,
                right=row[0].border.right,
                top=row[0].border.top,
                bottom=row[0].border.bottom
            )
            row[-1].border = Border(
                left=row[-1].border.left,
                right=ExcelStyle.ThickBorder,
                top=row[-1].border.top,
                bottom=row[-1].border.bottom
            )
        for col_top, col_bot in zip(rows[0], rows[-1]):
            col_top.border = Border(
                left=col_top.border.left,
                right=col_top.border.right,
                top=ExcelStyle.ThickBorder,
                bottom=col_top.border.bottom,
            )
            col_bot.border = Border(
                left=col_bot.border.left,
                right=col_bot.border.right,
                top=col_bot.border.top,
                bottom=ExcelStyle.ThickBorder,
            )
